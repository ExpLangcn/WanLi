from curses import beep
from distutils import core
import os
import sys
import json
import yaml
import time
import locale
import base64
import random
import requests
import datetime as d
import core.lib as lib
import core.lib as core
from openpyxl import Workbook
from rich.progress import track
from rich.console import Console
from rich.table import Column, Table
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()

console = Console()
sleeptime=random.randint(0, 1)

def fofa_info():
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['fofa_Key'] == '':
        core.error('FOFA','FOFA的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        attest = "https://fofa.info/api/v1/info/my?email=%s&key=%s"%(config['fofa_email'], config['fofa_Key'])
        r = requests.get(attest, verify=False)
        attest_json_data = json.loads(r.text)
        vip = attest_json_data['isvip']
        name = attest_json_data['username']
        Ver = attest_json_data['fofacli_ver']
        Level = attest_json_data['vip_level']
        core.correct('FOFA', '正在读取配置信息...')
        time.sleep(sleeptime)
        core.correct('FOFA', '正在验证账户信息...')
        time.sleep(sleeptime)
        if vip == False:
            core.correct('FOFA', 'Hi [boldblue]%s[/bold blue], 欢迎使用 [bold blue]%s[/bold blue], 当前FOFA-Cli版本为: [bold blue]%s[/bold blue], 您是FOFA的 [bold red]注册用户[/bold red], 抱歉您无法使    用API查询数据。'%(name,'WanLi Scan',Ver))
            sys.exit()
        else:
            if Level == 1:
                core.correct('FOFA', 'Hi [bold blue]%s[/bold blue], 欢迎使用 [bold blue]%s[/bold blue], 当前FOFA-Cli版本为: [bold blue]%s[/bold blue], 您是FOFA的 [bold red]普通会员[/bold red], 你可以  免费查询前 [bold red]100[/bold red] 条数据。'%(name,'WanLi Scan',Ver))
            elif Level == 2:
                core.correct('FOFA', 'Hi [bold blue]%s[/bold blue], 欢迎使用 [bold blue]%s[/bold blue], 当前FOFA-Cli版本为: [bold blue]%s[/bold blue], 您是FOFA的 [bold red]高级会员[/bold red], 你可以免费查询前 [bold red]10000[/bold red] 条数据。'%(name,'WanLi Scan',Ver))
            elif Level == 3:
                core.correct('FOFA', 'Hi [bold blue]%s[/bold blue], 欢迎使用 [bold blue]%s[/bold blue], 当前FOFA-Cli版本为: [bold blue]%s[/bold blue], 您是FOFA的 [bold red]企业会员[/bold red], 你可以免费查询前 [bold red]100000[/bold red] 条数据。'%(name,'WanLi Scan',Ver))

def fofa_search(keyword, date):
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['fofa_Key'] == '':
        core.error('FOFA','FOFA的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        fofa_info()
        console.print('''[[bold red]'''+ str(base64.b64decode("V2FuTGkgU2Nhbg=="), "utf-8") +'''[/bold red]] [[bold green]+[/bold green]] [bold green]FOFA[/bold green] > 当前要检测的语法为：[bold blue]%s[/bold blue]'''%(keyword))
        key = str(base64.b64encode(keyword.encode("utf-8")), "utf-8")
        url = "https://fofa.info/api/v1/search/all?email=%s&key=%s&qbase64=%s&size=%s&full=%s&fields=host,ip,port,title,domain"%(config['fofa_email'], config['fofa_Key'], key, config  ['fofa_limits'], config['Full'])
        r = requests.get(url, verify=False)
        json_data = json.loads(r.text)
        table = Table(show_header=True)
        table.add_column("ID", style="dim")
        table.add_column("IP")
        table.add_column("Port")
        table.add_column("Url")
        table.add_column("Domain")
        table.add_column("Title")
        console.print('''[[bold red]'''+ str(base64.b64decode("V2FuTGkgU2Nhbg=="), "utf-8") +'''[/bold red]] [[bold green]+[/bold green]] [bold green]FOFA[/bold green] > 查询语法：[bold blue]%s[/bold blue], 查询数据限制: [bold blue]%s[/bold blue], 成功查询到：[bold blue]%s[/bold blue] 条数据！'''%(keyword, config['fofa_limits'], json_data['size']))
        wb = Workbook()    #创建文件对象
        ws = wb.active
        ws['A1'] = 'IP'
        ws['B1'] = "Port"
        ws['C1'] = "Url"
        ws['D1'] = "Domain"
        ws['E1'] = "Title"
        file = open(config['Output'] + '/txt/' + date + "_keyword.txt",'a')
        for i in range(0,int(len(json_data['results']))):
            IP = json_data['results'][i][1]
            Port = json_data['results'][i][2]
            Url = json_data['results'][i][0]
            Title = json_data['results'][i][3]
            Domain = json_data['results'][i][4]
            file.write('%s\n'%(str(Url)))
            table.add_row(
                str(i+1),
                str(IP),
                str(Port),
                str(Url),
                str(Domain),
                str(Title)
                )
            ws.append([str(IP), str(Port), str(Url), str(Domain), str(Title)])

        for step in track(range(100)):
            time.sleep(0.015)
        console.print(table)
        wb.save(config['Output'] + "/xlsx/" + date + "_FOFA.xlsx")
        lib.info('The result file is saved at：' + config['Output'] + "/xlsx/" + date + "_FOFA.xlsx")

def fofa_domain(domain, dic):
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['fofa_Key'] == '':
        core.error('FOFA','FOFA的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        fofa_info()
        domain1 = 'domain="' + domain + '"'
        print(domain1)
        key = str(base64.b64encode(domain1.encode("utf-8")), "utf-8")
        url = "https://fofa.info/api/v1/search/all?email=%s&key=%s&qbase64=%s&size=%s&full=%s&fields=host"%(config['fofa_email'], config['fofa_Key'], key, config['fofa_limits'], config['Full'])
        r =requests.get(url, verify=False)
        json_data = json.loads(r.text)
        for i in range(0,int(len(json_data['results']))):
            dic.append(json_data['results'][i])

def quake_info():
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['quake_key'] == "":
        core.error('Quake', 'Quake的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        url = 'https://quake.360.cn/api/v3/user/info'
        headers = {
        "X-QuakeToken": config['quake_key']}
        r = requests.get(url=url, headers=headers)
        response = json.loads(r.text)
        name = response['data']['user']['fullname']
        credit = response['data']['credit']
        month_remaining_credit = response['data']['month_remaining_credit']
        core.correct('Quake','正在读取配置信息...')
        time.sleep(sleeptime)
        core.correct('Quake','正在验证账户信息...')
        time.sleep(sleeptime)
        console.print("[[bold red]"+ str(base64.b64decode("V2FuTGkgU2Nhbg=="), "utf-8") +"[/bold red]] [[bold green]+[/bold green]] [bold green]Quake[/bold green] > Hi [bold blue]%s[/bold blue], 欢迎使用 [bold blue]%s[/bold blue], 账户月度查询额度为:  [bold blue]%s[/bold blue], 账户月度剩余查询额度为 [bold blue]%s[/bold blue]。"%(name,'WanLi Scan',credit, month_remaining_credit))

def quake_search(keyword, date): # 实时数据查询
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['quake_key'] == "":
        core.error('Quake', 'Quake的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        quake_info()
        console.print('''[[bold red]'''+ str(base64.b64decode("V2FuTGkgU2Nhbg=="), "utf-8") +'''[/bold red]] [[bold green]+[/bold green]] [bold green]Quake[/bold green] >  当前要检测的语法为：[bold blue]%s[/bold blue]'''%(keyword))
        url = "https://quake.360.cn/api/v3/search/quake_service"
        headers = {
        "X-QuakeToken": config['quake_key'],
        "Content-Type": "application/json"
        }   
        data = {
         "query": keyword,
         "start": 0,
         "size": config['quake_limits'],
         "ignore_cache": config['ignore_cache'],
         "latest": config['latest'],
         "include": ["service.http.host", "ip", "port", "transport","service.http.title", "location.country_cn", "location.province_cn", "location.city_cn"],
         "start_time": config['start_time']
        }   
        r = requests.post(url="https://quake.360.cn/api/v3/search/quake_service", headers=headers, json=data)
        response = json.loads(r.text)
        console.print('''[[bold red]'''+ str(base64.b64decode("V2FuTGkgU2Nhbg=="), "utf-8") +'''[/bold red]] [[bold green]+[/bold green]] [bold green]Quake[/bold green] >  查询语 法：[bold blue]%s[/bold blue], 查询数据限制: [bold blue]%s[/bold blue]。'''%(keyword, config['quake_limits']))
        wb = Workbook()    #创建文件对象
        ws = wb.active
        ws['A1'] = 'IP'
        ws['B1'] = "Port"
        ws['C1'] = "TransPort"
        ws['D1'] = "Url"
        ws['E1'] = "Domain"
        ws['F1'] = "Title"
        ws['G1'] = "District Belong To"
        table = Table(show_header=True)
        table.add_column("ID", style="dim")
        table.add_column("IP")
        table.add_column("Port")
        table.add_column("TransPort")
        table.add_column("Url")
        table.add_column("Domain")
        table.add_column("Title")
        table.add_column("District Belong To")
        file = open(config['Output'] + '/txt/' + date + "_keyword.txt",'a')
        for i in range(0,int(len(response['data']))):
            data = response['data'][i]
            title = data['service']['http']['title']
            hostname = data['service']['http']['host']
            ip = data['ip']
            Port = data['port']
            transport = data['transport']
            url = "http://" + str(ip) + ":" + str(Port)
            country_cn_city_cn = data['location']['country_cn'] + data['location']['city_cn'] + data['location']['province_cn']
            file.write('%s\n'%(str(url)))
            table.add_row(
                str(i+1),
                str(ip),
                str(Port),
                str(transport),
                str(url),
                str(hostname),
                str(title),
                str(country_cn_city_cn),
            )
            ws.append([str(ip), str(Port), str(transport), str(url), str(hostname), str(title), str(country_cn_city_cn)])

        for step in track(range(100)):
            time.sleep(0.015)
        console.print(table)
        wb.save(config['Output'] + "/xlsx/" + date + "_Quake.xlsx")
        lib.info('The result file is saved at：' + config['Output'] + "/xlsx/" + date + "_Quake.xlsx")

def quake_domain(domain, dic):
    with open('config/config.yaml','r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.CLoader)
    if config['quake_key'] == "":
        core.error('Quake', 'Quake的KEY未在Config文件中配置，当前验证操作跳过。')
    else:
        quake_info()
        domain1 = 'domain:"' + domain + '"'
        url = "https://quake.360.cn/api/v3/search/quake_service"
        headers = {
        "X-QuakeToken": config['quake_key'],
        "Content-Type": "application/json"
        }   
        data = {
         "query": domain1,
         "start": 0,
         "size": config['quake_limits'],
         "ignore_cache": config['ignore_cache'],
         "latest": config['latest'],
         "include": ["service.http.host"],
         "start_time": config['start_time']
        }   
        r = requests.post(url="https://quake.360.cn/api/v3/search/quake_service", headers=headers, json=data)
        response = json.loads(r.text)
        for i in range(0,int(len(response['data']))):
            data = response['data'][i]
            dic.append(data['service']['http']['host'])
